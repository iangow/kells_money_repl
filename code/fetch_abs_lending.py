#!/usr/bin/env python3
"""
Fetch ABS Lending Indicators Table 3.

By default this fetches the March Quarter 2026 workbook used in the package.
Use --latest to fetch the current ABS latest-release workbook instead.
"""
import argparse
import hashlib
import os
import tempfile
import urllib.error
import urllib.request
import zipfile

import openpyxl


HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_OUT = os.path.join(HERE, "data_raw")
FILENAME = "560103.xlsx"
COMMITTED_RELEASE_URL = (
    "https://www.abs.gov.au/statistics/economy/finance/"
    "lending-indicators/mar-quarter-2026/560103.xlsx"
)
LATEST_RELEASE_URL = (
    "https://www.abs.gov.au/statistics/economy/finance/"
    "lending-indicators/latest-release/560103.xlsx"
)

REQUIRED_SERIES = {
    "Total dwellings excluding refinancing",
    "Purchase of existing dwellings",
    "Construction of dwellings",
    "Purchase of newly erected dwellings",
    "Alterations, additions and repairs",
    "External refinancing",
}


def sha16(data):
    return hashlib.sha256(data).hexdigest()[:16]


def fetch(url):
    req = urllib.request.Request(url, headers={"User-Agent": "kells-money-repl/0.1"})
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return resp.read(), resp.geturl()
    except urllib.error.URLError as exc:
        raise SystemExit(f"failed to fetch {url}: {exc}") from exc


def validate_workbook(filename, data):
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
        tmp.write(data)
        tmp.flush()
        if not zipfile.is_zipfile(tmp.name):
            raise SystemExit(f"{filename}: response is not an xlsx file")

        wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
        missing_sheets = {"Index", "Data1"} - set(wb.sheetnames)
        if missing_sheets:
            raise SystemExit(f"{filename}: missing sheets: {sorted(missing_sheets)}")

        rows = list(wb["Data1"].iter_rows(max_row=3, values_only=True))
        names, units, stype = rows[0], rows[1], rows[2]
        found = {
            str(names[i]).split(";")[3].strip()
            for i in range(1, len(names))
            if names[i]
            and len(str(names[i]).split(";")) > 3
            and str(units[i]) == "$ Millions"
            and str(stype[i]) == "Original"
        }
        missing = sorted(REQUIRED_SERIES - found)
        if missing:
            raise SystemExit(f"{filename}: missing required original $m series: {missing}")

        return wb.sheetnames


def write_bytes(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix=f".{os.path.basename(path)}.", dir=os.path.dirname(path))
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(data)
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)


def main():
    parser = argparse.ArgumentParser(description="Fetch ABS Lending Indicators Table 3.")
    parser.add_argument(
        "--output-dir",
        default=DEFAULT_OUT,
        help="Directory to write the fetched workbook. Defaults to data_raw/.",
    )
    parser.add_argument(
        "--latest",
        action="store_true",
        help="Fetch the current latest-release workbook instead of the committed March 2026 release.",
    )
    parser.add_argument(
        "--url",
        help="Override the ABS workbook URL. Useful for fetching a specific archived release.",
    )
    args = parser.parse_args()

    url = args.url or (LATEST_RELEASE_URL if args.latest else COMMITTED_RELEASE_URL)
    out_path = os.path.join(args.output_dir, FILENAME)
    old = open(out_path, "rb").read() if os.path.exists(out_path) else None

    data, final_url = fetch(url)
    sheets = validate_workbook(FILENAME, data)
    write_bytes(out_path, data)

    status = "new"
    if old is not None:
        status = "unchanged" if old == data else "changed"
    print(f"{FILENAME}: {status}; {len(data):,} bytes; sha256={sha16(data)}")
    print(f"source: {final_url}")
    print(f"sheets: {', '.join(sheets)}")


if __name__ == "__main__":
    main()
